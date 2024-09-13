import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from concurrent.futures import ThreadPoolExecutor, as_completed
import logging

# Constants
NUM_WORKERS = 10  # Adjust based on your system's capability

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),  # Logs to console
        logging.FileHandler("app.log")  # Logs to a file
    ]
)

def scrape_aku_result(reg_no, url_format):
    url = f"{url_format}{reg_no}"
    # Fetch the result page for the student
    logging.info(f"Fetching data for registration number: {reg_no}")
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'lxml')
    
    # Check if the result exists
    if "No Record Found" in response.text:
        logging.warning(f"No data found for registration number: {reg_no}")
        return None

    # Extract student details
    try:
        reg_no = soup.find(id="ctl00_ContentPlaceHolder1_DataList1_ctl00_RegistrationNoLabel").text.strip()
        student_name = soup.find(id="ctl00_ContentPlaceHolder1_DataList1_ctl00_StudentNameLabel").text.strip()
    except AttributeError:
        logging.error(f"Error parsing data for {reg_no}")
        return None

    # Extract theory marks
    theory_table = soup.find("table", id="ctl00_ContentPlaceHolder1_GridView1")
    theory_data = {}

    if theory_table:
        for row in theory_table.find_all("tr")[1:]:
            cells = row.find_all("td")
            subject_code = cells[0].text.strip()
            subject_name = cells[1].text.strip()
            theory_ese = cells[2].text.strip()
            theory_ia = cells[3].text.strip()
            theory_total = cells[4].text.strip()
            theory_grade = cells[5].text.strip()
            theory_credit = cells[6].text.strip()
            
            theory_data[subject_code] = {
                'Subject Name': subject_name,
                'Theory ESE': theory_ese,
                'Theory IA': theory_ia,
                'Theory Total': theory_total,
                'Theory Grade': theory_grade,
                'Theory Credit': theory_credit
            }

    # Extract practical marks
    practical_table = soup.find("table", id="ctl00_ContentPlaceHolder1_GridView2")
    practical_data = {}

    if practical_table:
        for row in practical_table.find_all("tr")[1:]:
            cells = row.find_all("td")
            subject_code = cells[0].text.strip()
            subject_name = cells[1].text.strip()
            practical_ese = cells[2].text.strip()
            practical_ia = cells[3].text.strip()
            practical_total = cells[4].text.strip()
            practical_grade = cells[5].text.strip()
            practical_credit = cells[6].text.strip()
            
            practical_data[subject_code] = {
                'Subject Name': subject_name,
                'Practical ESE': practical_ese,
                'Practical IA': practical_ia,
                'Practical Total': practical_total,
                'Practical Grade': practical_grade,
                'Practical Credit': practical_credit
            }

    # Extract CGPA
    cgpa_table = soup.find("table", id="ctl00_ContentPlaceHolder1_GridView3")
    cgpa = None

    sgpa = soup.find("span", id="ctl00_ContentPlaceHolder1_DataList5_ctl00_GROSSTHEORYTOTALLabel").text.strip()

    if cgpa_table:
        try:
            cgpa = cgpa_table.find_all("tr")[1].find_all("td")[-1].text.strip()
        except IndexError:
            logging.warning(f"CGPA data not found for registration number: {reg_no}")

    # Combine all data
    result = {
        'registration_number': reg_no,
        'student_name': student_name,
        'theory_marks': theory_data,
        'practical_marks': practical_data,
        'cgpa': cgpa,
        'sgpa':sgpa
    }
    
    return result


def scrape_beu_result(reg_no, url_format):
    url = f"{url_format}{reg_no}"
    
    # Fetch the result page for the student
    logging.info(f"Fetching data for registration number: {reg_no}")
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'lxml')
    
    # Check if the result exists
    if "No Record Found" in response.text:
        logging.warning(f"No data found for registration number: {reg_no}")
        return None

    # Extract student details
    try:
        reg_no = soup.find(id="ContentPlaceHolder1_DataList1_RegistrationNoLabel_0").text.strip()
        student_name = soup.find(id="ContentPlaceHolder1_DataList1_StudentNameLabel_0").text.strip()
    except AttributeError:
        logging.error(f"Error parsing data for {reg_no}")
        return None

    # Extract theory marks
    theory_table = soup.find("table", id="ContentPlaceHolder1_GridView1")
    theory_data = {}

    if theory_table:
        for row in theory_table.find_all("tr")[1:]:
            cells = row.find_all("td")
            subject_code = cells[0].text.strip()
            subject_name = cells[1].text.strip()
            theory_ese = cells[2].text.strip()
            theory_ia = cells[3].text.strip()
            theory_total = cells[4].text.strip()
            theory_grade = cells[5].text.strip()
            theory_credit = cells[6].text.strip()
            
            theory_data[subject_code] = {
                'Subject Name': subject_name,
                'Theory ESE': theory_ese,
                'Theory IA': theory_ia,
                'Theory Total': theory_total,
                'Theory Grade': theory_grade,
                'Theory Credit': theory_credit
            }

    # Extract practical marks
    practical_table = soup.find("table", id="ContentPlaceHolder1_GridView2")
    practical_data = {}

    if practical_table:
        for row in practical_table.find_all("tr")[1:]:
            cells = row.find_all("td")
            subject_code = cells[0].text.strip()
            if subject_code.endswith('P'):
                subject_code = subject_code[:-1]  # Remove the trailing 'P'
            subject_name = cells[1].text.strip()
            practical_ese = cells[2].text.strip()
            practical_ia = cells[3].text.strip()
            practical_total = cells[4].text.strip()
            practical_grade = cells[5].text.strip()
            practical_credit = cells[6].text.strip()
            
            practical_data[subject_code] = {
                'Subject Name': subject_name,
                'Practical ESE': practical_ese,
                'Practical IA': practical_ia,
                'Practical Total': practical_total,
                'Practical Grade': practical_grade,
                'Practical Credit': practical_credit
            }

    # Extract CGPA
    cgpa_table = soup.find("table", id="ContentPlaceHolder1_GridView3")
    cgpa = None

    sgpa = soup.find("span", id="ContentPlaceHolder1_DataList5_GROSSTHEORYTOTALLabel_0").text.strip()

    if cgpa_table:
        try:
            cgpa = cgpa_table.find_all("tr")[1].find_all("td")[-1].text.strip()
        except IndexError:
            logging.warning(f"CGPA data not found for registration number: {reg_no}")

    # Combine all data
    result = {
        'registration_number': reg_no,
        'student_name': student_name,
        'theory_marks': theory_data,
        'practical_marks': practical_data,
        'cgpa': cgpa,
        'sgpa': sgpa,
    }
    
    return result

def save_results_to_excel(url, reg_no_list, output_file):
    # Determine which scraping function to use
    if "akuexam" in url:
        scrape_function = scrape_aku_result
    else:
        scrape_function = scrape_beu_result

    results = []
    
    # Use ThreadPoolExecutor to fetch results concurrently
    with ThreadPoolExecutor(max_workers=NUM_WORKERS) as executor:
        future_to_reg_no = {executor.submit(scrape_function, reg_no, url): reg_no for reg_no in reg_no_list}
        
        for future in as_completed(future_to_reg_no):
            reg_no = future_to_reg_no[future]
            try:
                result = future.result()
                if result:
                    results.append(result)
            except Exception as e:
                logging.error(f"Error processing registration number {reg_no}: {e}")
    
    # Save the results to Excel
    wb = Workbook()
    
    for result in results:
        for subject_code in result['theory_marks']:
            sheet_name = result['theory_marks'][subject_code]['Subject Name']
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(title=sheet_name)  # Truncate sheet name if needed
                # Write the header
                ws.append([
                    "Registration No", "Student Name",
                    "Subject Code", "Subject Name",
                    "Theory ESE", "Theory IA", "Theory Total", "Theory Grade", "Theory Credit",
                    "Practical ESE", "Practical IA", "Practical Total", "Practical Grade", "Practical Credit",
                    "SGPA","CGPA"
                ])
                
            ws = wb[sheet_name]
            theory = result['theory_marks'].get(subject_code, {})
            practical = result['practical_marks'].get(subject_code, {})
            
            row = [
                result['registration_number'],
                result['student_name'],
                subject_code,
                theory.get('Subject Name', practical.get('Subject Name', '')),
                theory.get('Theory ESE', ''),
                theory.get('Theory IA', ''),
                theory.get('Theory Total', ''),
                theory.get('Theory Grade', ''),
                theory.get('Theory Credit', ''),
                practical.get('Practical ESE', ''),
                practical.get('Practical IA', ''),
                practical.get('Practical Total', ''),
                practical.get('Practical Grade', ''),
                practical.get('Practical Credit', ''),
                result['sgpa'],
                result['cgpa']
            ]
            ws.append(row)
            logging.info(f"Stored data for registration number: {result['registration_number']} in sheet: {sheet_name}")

    # Remove the default sheet if other sheets are present
    if len(wb.sheetnames) > 1 and 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])  # Remove the default sheet created by Workbook()
    
    # Save the workbook
    wb.save(output_file)
    logging.info(f"Data saved to {output_file}")


# Example usage
# url = "https://results.akuexam.net/ResultsBTechBPharm1stSemPub2021.aspx?Sem=I&RegNo="
# reg_no_list = [f"{i:02d}" for i in range(1, 11)]  # Example registration numbers
# output_file = "results.xlsx"

# scrape_and_save_results(url, reg_no_list, output_file)